from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from collections import defaultdict
from datetime import date

# 创建一个 excel 文件
wb = Workbook()
n = input('大概多少个款项录入? ')
if n=='':
    n=10
else:
    n = int(n)

# 今日日期
today = date.today()
# 年月日
date = today.strftime("%m%d")

# 创建一个你自己命名的表格
wbname = input ("你想用的文件名?:")  # 自己填的信息
if wbname=='':
    wbname='yj'
filename = "输入表-"+date+wbname+".xlsx"

# 转入当前活跃的表格
ws = wb.active
ws.title="输入表"

group_val = DataValidation(type="list",formula1='"固定-初始版,固定-调整版,变动-调整版"',showErrorMessage=False,allow_blank=True)
ws.add_data_validation(group_val)

#创建表头
ws['A1']='类型'
ws['B1'] = '机构'
ws['C1']='科目'
ws['D1']='金额'

# 创建下拉菜单
for i in range(2,2+n):
    group_val.add(ws["A"+str(i)])

wb.save(filename)
