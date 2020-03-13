from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import date
import glob

# 加载科目表,成本中心表
mb = load_workbook('自动生成预算导入模板.xlsx')
km = mb['科目表']
km_acName = defaultdict(str)
km_acVal =defaultdict(str)
for i in range(2,35):
    km_acName[km['B'+str(i)].value]=km['A'+str(i)].value
    km_acVal[km['A'+str(i)].value]=km['B'+str(i)].value
    
cbzx = mb['成本中心']
cbzx_acName = defaultdict(str)
cbzx_acVal =defaultdict(str)
for i in range(3,199):
    cbzx_acName[cbzx['B'+str(i)].value]=cbzx['A'+str(i)].value
    cbzx_acVal[cbzx['A'+str(i)].value]=cbzx['B'+str(i)].value

# 今日日期
today = date.today()
# 月日
date = today.strftime("%m%d")

# 创建一个 excel 文件
wb = Workbook()

# 转入当前活跃的表格
ws = wb.active
ws.title="明细账"

# 创建一个你自己命名的表格
wbname = input ("你想用的文件名?:")  # 自己填的信息
if wbname=='':
    wbname='yj'
bzstr = input ("备注要用到的信息? 比如，‘日常费用’:")
if bzstr=='':
    bzstr='日常费用'
filename = "预算明细账模板-"+date+wbname+".xlsx"

# 关于成本中心，金额的信息录入
info = []
cur = '1'
group={'1':'固定-初始版','2':'固定-调整版','3':'变动-调整版'}
print('请输入新的款项，格式为：类别 机构 科目 金额（注意空格）')
print('类别有3个选项，分别输入对应数字即可（1-固定-初始版,2-固定-调整版,3-变动-调整版）')
while cur!='':
    cur = input('开始录入：')
    templst = cur.split()
    if 0<len(templst)<4:
        print('上一条信息不足，请分别输入四个字符')
        cur = input('重新录入：')
    elif len(templst)==4:
        info.append((group[templst[0]],templst[1],templst[2],float(templst[3])))


# 有多少条信息？
num_entry = len(info)
print('你成功录入了{}个款项'.format(num_entry))

#创建表头
ws['A1']='类型'
ws['B1'] = '年'
for i in range(67,67+12):
    ws[chr(i)+'1']=str(i-66)+"月"
ws['O1']='备注'
ws['P1']='组织机构'
ws['Q1']='预算科目编码'
ws['R1']='预算科目'

# 录入信息
month = int(today.strftime("%m"))
year = int(today.strftime("%Y"))

for i,e in enumerate(info):
    ws['A'+str(i+2)]=e[0]
    ws['B'+str(i+2)]=year
    # 金额信息
    ws[chr(66+month)+str(i+2)]=e[3]
    # 备注信息
    ws['O'+str(i+2)]='预拨'+str(month)+'月'+e[0]+bzstr+date+'-yj'
    # 成本中心
    ws['S'+str(i+2)]=e[1]
    ws['P'+str(i+2)]=cbzx_acName[e[1]]
    # 预算科目
    ws['R'+str(i+2)]=e[2]
    ws['Q'+str(i+2)]=km_acName[e[2]]

# 改变列宽
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter # Get the column name
    for cell in col:
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    if max_length>0:
        adjusted_width = max_length*1.5
        ws.column_dimensions[column].width = adjusted_width

ws.column_dimensions['A'].width*= 1.25
ws.column_dimensions['O'].width*= 1.2
ws.column_dimensions['R'].width*= 1.5
ws.column_dimensions['S'].width*= 1.5

# 保存文件
wb.save(filename)
